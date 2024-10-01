# -*- coding: utf-8 -*-
"""
Created on Wed May 29 18:05:16 2024

@author: Ion
"""
import pandas as pd
import json
debug = 1
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def write_excel(test_pd,debug = 0 ):
    from datetime import datetime
    today =  datetime.today().strftime('%Y-%m-%d_%H_%M')


    modo = {0:'niente', 1:'base' , 2 : 'arlecchino'}
    df = pd.concat(test_pd)
    if debug != 0:
        with pd.ExcelWriter(f'check_{today}_{modo[debug]}.xlsx', engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='recap')
            workbook = writer.book
            worksheet = writer.sheets['recap']
        
            # Define the fills
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            if shifts == 3:
                columns_to_color = [5, 7, 9, 11, 13, 15]
            if shifts == 1:
                columns_to_color = [5, 7]
            for row in range(2, len(df) + 2):  # DataFrame rows start from row 2 in Excel
                if debug ==1:
                    cell = worksheet.cell(row=row, column=1)
                    if cell.value == 'RECAP':
                        for col in columns_to_color:
                            
                            cell = worksheet.cell(row=row, column=col)
                            
                            if cell.value != '':
                                if float(cell.value) < 1:
                                    cell.fill = green_fill
                                elif  float(cell.value) > 2:
                                    cell.fill = red_fill
                                elif  float(cell.value) > 1:
                                    cell.fill = yellow_fill
                if debug == 2:
                    for col in columns_to_color:
                        
                        cell = worksheet.cell(row=row, column=col)
                        
                        if cell.value != '':
                            if float(cell.value) < 1:
                                cell.fill = green_fill
                            elif  float(cell.value) > 2:
                                cell.fill = red_fill
                            elif  float(cell.value) > 1:
                                cell.fill = yellow_fill
        
def exception_IKM11001(differenza,machine,checking_plc_or_mes ,shift_report_machine,  site_id , PBI_DATA_machine , Type , shifts , this_issues,general_perc,  debug =0 ):
    this = {}
    this_diff = 0
    this['Machine'] = machine
    this['site_id'] = site_id
    this['type'] = Type
    shifts_of_this_machine = [0,1,3]
    problema = 0
    try:
        for what in list(shift_report_machine.keys()):
            for valore in shifts_of_this_machine:
                if PBI_DATA_machine != []:
                    
                    machine_PBI = [i for i in PBI_DATA_machine  if int(i['SHIFT_NUMBER']) == valore+1][0][checking_plc_or_mes[what]]
                    
                    if valore ==3 : 
                        problema = 1
                        machine_SR = shift_report_machine[what][2]
                    else:
                        machine_SR = shift_report_machine[what][valore]
                        
                        
                    if machine_PBI == '' :
                        machine_PBI = 0
                    else:
                        machine_PBI = int(machine_PBI)
                    
                    diff = machine_PBI - machine_SR
                    differenza += diff
                    if machine_PBI != 0 :
                        this_perc = diff/machine_PBI*100  
                        general_perc.append({'shift' : valore+1 , 'perc' : this_perc, 'pcl_mes' : what})
                    else: this_perc = 0
                    if machine_PBI == 0 and machine_SR == 0: 
                        diff = ""
                        this_perc = ""
                else:
                    print(f'no data for {machine} , from Shift Report was {shift_report_machine}')
                if problema ==1: 
                    valore -=1
                this[f'Shift {valore+1} {what}'] = diff
                this[f'Shift {valore+1} {what} % diff'] = str(this_perc)
                #print("ok for machine strange")
    except :
        if debug == 1:
            print(f'issues on {machine} on {file} , Type : {Type} ,checking_plc_or_mes = {checking_plc_or_mes[what]}  ')
        this_issues.append(machine)
        this[f'Shift {valore+1} {what}'] = 'issue'
        #this[f'Shift {valore+1} descr'] = 'issue'
        this[f'Shift {valore+1} % diff'] = 'issue'
    return this , this_issues , general_perc


def check_this_machine(differenza,machine,checking_plc_or_mes ,shift_report_machine,  site_id , PBI_DATA_machine , Type , shifts , this_issues, general_perc , debug =0 ):
    this = {}
    this_diff = 0
    this['Machine'] = machine
    this['site_id'] = site_id
    this['type'] = Type
    try:
        
        for what in list(shift_report_machine.keys()):
            for valore in range(shifts):
                if PBI_DATA_machine != []:
                    
                    machine_PBI = [i for i in PBI_DATA_machine  if int(i['SHIFT_NUMBER']) == valore+1][0][checking_plc_or_mes[what]]
                    machine_SR = shift_report_machine[what][valore]
                    if machine_PBI == '' :
                        machine_PBI = 0
                    else:
                        machine_PBI = int(machine_PBI)
                    
                    diff = machine_PBI - machine_SR
                    differenza += diff
                    if machine_PBI != 0 :
                        this_perc = diff/machine_PBI*100  
                        this_perc = round(this_perc,3)
                        general_perc.append({'shift' : valore+1 , 'perc' : this_perc, 'pcl_mes' : what ,'Type' : Type})
                    else: this_perc = 0
                    if machine_PBI == 0 and machine_SR == 0: 
                        diff = ""
                        this_perc = ""
                else:
                    print(f'no data for {machine} , from Shift Report was {shift_report_machine}')
                #print(f'Per la {machine} , turno {valore+1} , abbiamo PBI : {machine_PBI} , SR : {machine_SR} , diff = {diff}')
                this[f'Shift {valore+1} {what}'] = diff
                #this[f'Shift {valore+1} {what} descr'] = f'{machine_PBI} - {str(machine_SR)}  = {diff} '
                this[f'Shift {valore+1} {what} % diff'] = str(this_perc)
    except :
        if debug == 1:
            print(f'issues on {machine} on {file} , Type : {Type} ,checking_plc_or_mes = {checking_plc_or_mes[what]}  ')
        this_issues.append(machine)
        this[f'Shift {valore+1} {what}'] = 'issue'
        #this[f'Shift {valore+1} descr'] = 'issue'
        this[f'Shift {valore+1} % diff'] = 'issue'
    return this , this_issues , general_perc

def do_recap(plc_or_mes_perc,plant_recap,debug = 0):
    plc_or_mes_perc = [i for i in plc_or_mes_perc if i!= []]
    for valore in range(shifts):
        for what in list(shift_report_machine.keys()):
            plant_recap[f'Shift {valore+1} {what}'] = ''
            plant_recap[f'Shift {valore+1} {what} % diff'] = ''
            perc_gen = []
            for perc_detail in plc_or_mes_perc:
                this = [i for i in perc_detail if i['shift'] == valore+1 and i['pcl_mes'] == what ]
                if this != []:
                    perc_gen.append(this)
            if perc_gen != []:
                percent = [i[0]['perc'] for i in perc_gen]
                if debug ==1:
                    plant_recap[f'Shift {valore+1} {what}'] = f'{percent.count(0.0)}  ok / {len(percent)} total ,{round(percent.count(0.0) /len(percent) * 100 ,2) } %  '
                else:
                    plant_recap[f'Shift {valore+1} {what}'] = f' {len(percent)}'
                plant_recap[f'Shift {valore+1} {what} % diff'] = round( sum(percent)/len(percent)  ,2)
    return plant_recap
    


f = open('data.csv' ,'r').read().splitlines()
f = [i.split(',') for i in f]
#keys = ['SHIFT_START' ,'SHIFT_NUMBER', 'SITE_ID' , 'WORK_CENTER' , 'DECPROD' , 'DECSCRAP' , 'PLCPROD' , 'PLCSCRAP' ,'WORK_ORDER'  ]
keys = ['SHIFT_START','SHIFT_NUMBER' , 'SITE_ID' , 'WORK_CENTER' , 'DECPROD' , 'DECSCRAP' , 'PLCPROD' , 'PLCSCRAP'  ]
differenze = {}
f = f[1:]
PBI_DATA = [{keys[i]: ff[i] for i in range(len(keys))} for ff in f]
PBI_DATA_pd  = pd.DataFrame( data = PBI_DATA, columns= [a for a in PBI_DATA[0].keys()])
PBI_DATA_plants = list(set([i['SITE_ID'] for i in PBI_DATA]))

MAP_PLANTS = {'FRE':1661 , 'CHA' : 1105 , 'CHI' : 1161, 'MEN' : 1131 , 'CHA': 1105 , 'PES': 1162 ,'LNB' : 1101 , 'CKY' : 1061,'MID':1261}
file_check = {a: int(x) for x in PBI_DATA_plants  for a in MAP_PLANTS if MAP_PLANTS[a] == int(x)}
recap_of_recaps = []
issues = {}
shifts = 3 # how many shifts ? this will go from 1 to 3 indicating how many shits to check 
all_perc = []
#for every plant 
files =  list(file_check.keys())
to_remove =  ['CHI', 'CHA','FRE','LNB']
for file in to_remove :
    pass
    #files.remove(file)
print(files)
for file in files:
    print(f"doing {file}")
    SHIFT_REPORT_RECAP_THIS_PLANT = json.loads(open(f'{file}.txt' , 'r').read().splitlines()[0])
    PBI_DATA_THIS_PLANT  = [i for i in PBI_DATA if i['SITE_ID'] == str(file_check[file])]
    frek = list(SHIFT_REPORT_RECAP_THIS_PLANT.keys())
    plc = [i for i in frek if 'plc/time' in i]
    mes = [i for i in frek if 'mes/time' in i]
    PLC_SHIFT_REPORT_RECAP_THIS_PLANT = SHIFT_REPORT_RECAP_THIS_PLANT[plc[0]]
    MES_SHIFT_REPORT_RECAP_THIS_PLANT = SHIFT_REPORT_RECAP_THIS_PLANT[mes[0]]
    differenza = 0 
    general_check = []
    SHIFT_REPORT_OF_PC_OR_MES = [PLC_SHIFT_REPORT_RECAP_THIS_PLANT,MES_SHIFT_REPORT_RECAP_THIS_PLANT ]
    
    this_issues = []
    plant_perc = []
    for choose in range(len(SHIFT_REPORT_OF_PC_OR_MES)) :  #ciclo sia sui valori del plc che su quelli del mes
        plc_or_mes_perc = []
        recap = []
        plc_or_mes_check = {0 : 'plc' , 1 : 'mes'}
        if choose == 0:
            checking_plc_or_mes = {'production_produced' :'PLCPROD' ,'scraped' : 'PLCSCRAP'  }
        else:
            checking_plc_or_mes = {'production_produced' :'DECPROD' ,'scraped' : 'DECSCRAP'  }
            
            
        site_id = str(file_check[file])
        Type =  plc_or_mes_check[choose]
        plant_recap = {'Machine' : 'RECAP' , 'site_id' : site_id ,'type' : Type }
        
        for machine in SHIFT_REPORT_OF_PC_OR_MES[choose] : 
            
            PBI_DATA_machine = [i for i in PBI_DATA_THIS_PLANT if i['WORK_CENTER'] == machine]
            shift_report_machine = SHIFT_REPORT_OF_PC_OR_MES[choose][machine]
            if [int(i['SHIFT_NUMBER']) for i in PBI_DATA_machine ] != [1,2,4] : 
                this , this_issues  , general_perc = check_this_machine(differenza, machine , checking_plc_or_mes,shift_report_machine , site_id ,PBI_DATA_machine ,  Type ,shifts , this_issues,  [] ,  debug)
            else: 
                print('entering exeption')
                this , this_issues , general_perc = exception_IKM11001(differenza, machine , checking_plc_or_mes,shift_report_machine , site_id ,PBI_DATA_machine ,  Type ,shifts , this_issues,[] ,    debug)
            recap.append(this)
           ######################## 
            plc_or_mes_perc.append(general_perc)
            plant_perc.append(general_perc) ############
            ########################
            
        this_plant_recap = do_recap(plc_or_mes_perc , plant_recap)
        recap.append(this_plant_recap)
        all_perc.append(this_plant_recap)
        general_check.append({plc_or_mes_check[choose] : recap})
    
    plc_recap_pd = pd.DataFrame( data = general_check[0]['plc'], columns= [a for a in general_check[0]['plc'][0].keys()])
    mes_recap_pd = pd.DataFrame( data = general_check[1]['mes'], columns= [a for a in general_check[1]['mes'][0].keys()])
    recap_of_recaps.append(general_check[0]['plc'])
    recap_of_recaps.append(general_check[1]['mes'])
    if this_issues != []:
        print(f'on {file} found {len(this_issues)} issues')
        issues[file_check[file]] = this_issues








test_pd = [pd.DataFrame(data = i, columns= [a for a in i[0].keys()]) for i in recap_of_recaps]




write_excel(test_pd,debug)
write_excel(test_pd,2)


if debug ==1 :
    file = 'MEN'
    choose = 0
    choose = 0

    what = 'production_produced'
    #PL = issues[1131]
    #machine = PL[0]