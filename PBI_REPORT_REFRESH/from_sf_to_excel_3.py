# -*- coding: utf-8 -*-
"""
Created on Tue Aug 13 10:00:38 2024

@author: Ion
"""
# -*- coding: utf-8 -*-
from datetime import datetime, timedelta

time_windows = {'euro' : [[2,6],[10,13],[18,23]] , 'ww' : [[20,5],[5,13],[13,20]]}
def generate_time_ranges(times):
    start_date = datetime(2024, 7, 31)
    end_date = datetime(2024, 8, 30)
    delta = timedelta(days=1)
    
    time_ranges = []
    current_date = start_date
    
    while current_date <= end_date:
        for x in times :
        # Time range 1: 2 AM to 6 AM
            if x[1] > x[0]:
                time_range = (current_date.replace(hour=x[0], minute=30), current_date.replace(hour=x[1], minute=30))
            else:
                time_range = (current_date.replace(hour=x[0], minute=30),  (current_date + timedelta(days=1)).replace(hour=x[1], minute=30))
            if   time_range[0] >=  datetime(2024, 7, 31, 20)    : 
                time_ranges.append(time_range)
        current_date += delta
    
    return time_ranges

# Example usage
time_ranges = generate_time_ranges(time_windows['ww'])
time_ranges = sorted(time_ranges, key=lambda x: x[0])

#╗for start, end in time_ranges:
#   print(f"Start: {start}, End: {end}")


import json
from datetime import datetime  
import pytz
import pandas as pd

amount_of_time = 15
cf = json.load(open('new_config.json','r'))
local_tz = pytz.timezone(cf['local_tz'])
utc = pytz.timezone('UTC')
euro = cf['euro']
ww = cf['ww']
xl = cf['xl']
Month = 'Agosto'



def get_lista_plant(f,forma):
    dizionari = []
    for i in f:
        diz = {}
        diz['plant'] = int(i[2])
        diz['success'] = int(i[3])
        diz['from'] =  datetime.strptime( i[5] , forma)
        diz['to'] =    datetime.strptime( i[6] , forma)
        diz['start'] = datetime.strptime( i[7] , forma)
        diz['end'] =   datetime.strptime( i[8] , forma)
        diz['diff'] = diz['start'] - diz['to'] 
        diz['diff'] = int(diz['diff'].seconds)//60
        dizionari.append(diz) 
    return dizionari
def select_world_zone(lista_plants):
    which_plants = 'null'
    key_0 = list(set([i['plant'] for i in lista_plants]))[0]
    if key_0 in euro:
        which_plants = 'euro'
    else:
        which_plants = 'ww'
    return which_plants

amount_of_time = 20
forma = '%d/%m/%Y, %H:%M:%S'
f = open(f'REFRESH/{Month}.txt','r').read().splitlines()
f = [i.split('\t') for i in f]
refreshes = [[i[2],i[3]] for i in f if i[4] == 'Completata']
refreshes = [[datetime.strptime(i[0], forma), datetime.strptime(i[1], forma)] for i in refreshes]
refreshes = refreshes[::-1]
##AT THIS POINT I HAVE THE PBI REFRESHSES, IN UTC+1 OR LOCAL TIME


forma = '%Y-%m-%d %H:%M:%S.%f'
SF_data = open(f'SF_{Month}.txt','r').read().splitlines()
SF_data = [i.split('\t') for i in SF_data]

lista_plants = get_lista_plant(SF_data,forma)
lista_pd = pd.DataFrame(data = lista_plants,columns= [i for i in lista_plants[0].keys()])
combo_data = {}


from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
yellow = "00FFFF00"
PBI_yellow = "ffffe199"
SF_Yellow = "ffffffcc"
workbook = load_workbook(filename=cf['xl'])
sheet =  workbook[f"{Month}"]



for world in [euro, ww]:
    import datetime

    
    if world == euro:
        start_row = 3
        for x in range(len(refreshes)) : 
            refreshes[x][0] = refreshes[x][0] - datetime.timedelta(hours=2)
            refreshes[x][1] = refreshes[x][1] - datetime.timedelta(hours=2)
    else:
        start_row = 21
        
    world_data = [i for i in lista_plants if i['plant'] in world]
    from datetime import datetime  

    
    if world_data != []: 
        world_zone = select_world_zone(world_data)
        time_ranges = generate_time_ranges(time_windows[world_zone])
        time_ranges = sorted(time_ranges, key=lambda x: x[0])
        time_ranges = [i for i in time_ranges if i[0] < datetime.now()] #fino a oggi
        c = 0
        column_shift = 4
        
        for time_range in time_ranges:
            c+=1
            if c== 3:
                c=0
            wor = world
            print(f'doing {time_range}')
            returning = [{'plants':i} for i in world ]
            this_times = [w for w in world_data if w['start'] > time_range[0] and w['end'] < time_range[1] ]  #### dfdoes this consider recalculations?
            
            
            this_times_pd = pd.DataFrame(data = this_times,columns= [i for i in this_times[0].keys()])
            if len([i['plant'] for i in this_times]) != len(set([i['plant'] for i in this_times])):
                diff = [i for i in this_times if [i['plant'] for i in this_times].count(i['plant']) >1 ]
                print(f'found duplicates records on {time_range} for {diff} \n \n \n total of {len(diff)} records \n \n ')
                
                discard = [i['diff'] for i in diff]
                ind = discard.index(max(discard))
                this_to_discard = diff[ind]
                this_times = [i for i in this_times if i!=this_to_discard ]
                this_times_pd_2 = pd.DataFrame(data = this_times,columns= [i for i in this_times[0].keys()])
            if len(this_times) !=  len(world)  and world == ww :
                wor =  [1011, 1261, 1361, 1822, 1823, 1824, 1891,2008 ,  2551]
                
                import datetime
                new_end_time = time_range[1] + datetime.timedelta(minutes=30)
                time_range = (time_range[0] , new_end_time)
                from datetime import datetime  

                this_times = [w for w in world_data if w['start'] > time_range[0] and w['end'] < time_range[1] ]  #### dfdoes this consider recalculations?
                wor = [i['plant'] for i in this_times]
                this_times = [i for i in this_times if i['plant'] in wor ]
                
                miss = [i for i in world if i not in [i['plant'] for i in this_times]]
            calculation_end = [[this_times[ind]['plant'],this_times[ind]['end']] for ind in range(len(wor)) ]
            
            for i in range(len(returning)):
                if returning[i]['plants'] in wor:
                    plant = [tt for tt in this_times if returning[i]['plants'] == tt['plant']][0]
                    returning[i]['time'] = plant['end']
                    if plant['diff'] > amount_of_time and plant['plant'] not in [2222]:
                        returning[i]['time'] = 'the calc took place ' + str(plant['diff']) + ' minutes after shift end!' 
                        returning[i]['refresh'] =' error'
                    if plant['from'] == plant['to']:
                        returning[i]['time'] = 'start = end!'
                        returning[i]['refresh'] =' error'
                    if plant['success'] == 0:
                        returning[i]['time'] = 'Success = FALSE!'
                        returning[i]['refresh'] =' error'
                    if type(returning[i]['time']) != str:
                        for ref in range(len(refreshes)-1):
                            if plant['end'] > refreshes[ref][0] and plant['end'] <= refreshes[ref+1][0]:
                                returning[i]['refresh'] = refreshes[ref+1][1]
                                time_diff = returning[i]['refresh'] - returning[i]['time']
                                time_diff_hours = time_diff.seconds/3600
                                if time_diff_hours > 3 :
                                    returning[i]['refresh'] = 'no refresh'
                        
                elif returning[i]['plants'] not in wor and wor != world and c == 0: #I ignore multiple shifts only when it's the night shift with few sites having it .
                    returning[i]['time'] = 'ignore'
                    returning[i]['refresh'] = 'ignore'
                else:
                    returning[i]['time'] = 'no time'
                    returning[i]['refresh'] = 'no refresh'
            try:
                if world_zone != 'euro' :
                    returning =  [[i['plants'],i['time'].astimezone(pytz.timezone('Europe/Berlin')),i['refresh'].astimezone(pytz.timezone('Europe/Berlin'))] if type(i['time']) != str else [i['plants'],i['time'],i['refresh']] for i in returning ]
                    
            except:
                    print(f"issues here in {returning}")
                    returning = [[i['plants'],i['time'].astimezone(pytz.timezone('Europe/Berlin'))] if type(i['time']) != str  else [i['plants'],i['time']] for i in returning ]
                    
            try:
                if world_zone == 'euro' : #from OCT to MARCh, daylight saving time = 'Asia/Famagusta' /      UTC+2 : Asia/Baku
                    returning = [[i['plants'],i['time'].astimezone(pytz.timezone('Asia/Baku')),i['refresh'].astimezone(pytz.timezone('Asia/Baku'))] if type(i['time']) != str  else [i['plants'],i['time'],i['refresh']] for i in returning ]
            except:
                    print(f"issues here in {returning}")
                    returning = [[i['plants'],i['time'].astimezone(pytz.timezone('Asia/Baku'))] if type(i['time']) != str  else [i['plants'],i['time']] for i in returning ]

            ret = ''
            combo_data[time_range] = returning
            
            
            
            for x in range(len(returning)):
                if len(returning[x])<=2 and returning[x][1] != 'ignore':
                    if type(returning[x][1]) != str :
                        what_to_write = returning[x][1].replace(tzinfo=None)
                        
                    else:
                        what_to_write = returning[x][1]
                        sheet.cell(row = start_row+x, column=column_shift).fill =  PatternFill(start_color=yellow, end_color=yellow, fill_type = "solid")
                        sheet.cell(row = start_row+x, column=column_shift+1).fill =  PatternFill(start_color=yellow, end_color=yellow, fill_type = "solid")
                        sheet.cell(row = start_row+x, column=column_shift+2).fill =  PatternFill(start_color=yellow, end_color=yellow, fill_type = "solid")
                    sheet.cell(row = start_row+x, column=column_shift).value = what_to_write
                    sheet.cell(row = start_row+x, column=column_shift+2).value = ' '
                    
                elif len(returning[x])<=2 and returning[x][1] == 'ignore':
                    pass
                    
                elif returning[x][2] == 'ignore' :
                    pass
                elif returning[x][2] not in  (' error' , 'no refresh', 'ignore') :
                    sheet.cell(row = start_row+x, column=column_shift).value = returning[x][1].replace(tzinfo=None)
                    sheet.cell(row = start_row+x, column=column_shift+1).value = returning[x][2].replace(tzinfo=None)
                    #highlight delays in the Report calculation 
                    this_row = [i for i in this_times if i['plant'] == returning[x][0]][0]
                    this_td = this_row['end'] - this_row['to']
                    this_td_min = this_td.seconds / 60
                    if this_td_min > 45:
                        #highlith when a shift took more than 45 min after the shift end to be processed
                        sheet.cell(row = start_row+x, column=column_shift).fill =  PatternFill(start_color=SF_Yellow, end_color=SF_Yellow, fill_type = "solid")
                    if (returning[x][2]-returning[x][1]).seconds/60 > 45:
                        #highlith when a shift took more than 45 min to be available in PBI
                        sheet.cell(row = start_row+x, column=column_shift+1).fill =  PatternFill(start_color=PBI_yellow, end_color=SF_Yellow, fill_type = "solid")
                        sheet.cell(row = start_row+x, column=column_shift+2).fill =  PatternFill(start_color=PBI_yellow, end_color=SF_Yellow, fill_type = "solid")
                        
                else:
                    sheet.cell(row = start_row+x, column=column_shift).value = returning[x][1]
                    sheet.cell(row = start_row+x, column=column_shift+1).value = returning[x][2]
                    
                    sheet.cell(row = start_row+x, column=column_shift).fill =  PatternFill(start_color=yellow, end_color=yellow, fill_type = "solid")
                    sheet.cell(row = start_row+x, column=column_shift+1).fill =  PatternFill(start_color=yellow, end_color=yellow, fill_type = "solid")
                    sheet.cell(row = start_row+x, column=column_shift+2).fill =  PatternFill(start_color=yellow, end_color=yellow, fill_type = "solid")

                    
            column_shift+= 3 
            if c == 0: 
                column_shift+=1


workbook.save(filename=f'test_out_{Month}.xlsx')
"""
from openpyxl import Workbook
from openpyxl import load_workbook

workbook = load_workbook(filename=cf['xl'])




sheet =  workbook["Luglio"]
#sheet.cell(row = start_row, column=4).value
start= 3

for x in range(len(returning)):
    
    sheet.cell(row=start+x, column=4).value = returning[x][1].replace(tzinfo=None)
    sheet.cell(row=start+x, column=5).value = returning[x][2].replace(tzinfo=None)

workbook.save(filename=cf['xl'])

"""




















